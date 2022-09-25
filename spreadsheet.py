import openpyxl as xl
from openpyxl.chart import BarChart3D, Reference


#Sheet1
wb = xl.load_workbook('transactions.xlsx')

#Access Sheet1(Sheet is Upper case sensitive)
sheet= wb['Sheet1']

#To get our Cell
cell = sheet['a1']
#    OR
# cell = sheet.cell(1, 1)
# print(cell.value)

#To get how may rows we have in a Sheet
# print (sheet.max_row)

#To Generate the numbers in a rows
# for row in range(1, sheet.max_row + 1):
#     print(row)

# To get the values in any column (third column)
# change one to two to ommit row one (Price)
# for row in range(2, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     print(cell.value)

# How to correct data in the Sheet
# for row in range(2, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     corrected_price = cell.value * 0.9
#     print(corrected_price)

# Adding the corrected_price to a new row
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Adding a Chart to the Sheet
# import BarChart and Reference in line 2
data = Reference(
    sheet, min_row=2,
    max_row=sheet.max_row,
    min_col=4, max_col=4,
)

chart = BarChart3D()
chart.add_data(data)
sheet.add_chart(chart, 'e2')
s = chart.series[0]
s.graphicalProperties.line.solidFill="999"
s.graphicalProperties.solidFill="ff9900"

# We Save
wb.save('New_Work.xlsx')
print("Successfully Saved!")
